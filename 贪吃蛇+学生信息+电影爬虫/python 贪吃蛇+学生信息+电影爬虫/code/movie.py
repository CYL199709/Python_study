import pymysql
import datetime
from tkinter import *

import tkinter
from  tkinter  import ttk

import requests
import os
from bs4 import BeautifulSoup
import openpyxl
import threading
import matplotlib.pylab as plt
import re
import matplotlib.pyplot as plt
import numpy as np
os.chdir('D:\爬虫数据')
conn = pymysql.connect('180.76.177.224','root','root','douban')
cursor = conn.cursor()
def get_movie_top250_name(soup):
    targets = soup.find_all('span', class_="title")  # 用BeautifulSoup找寻一个内容为一个列表
    targets_name = re.findall(r'.*?title">(.*?)<\/span', str(targets))  # 用正则表达式去掉标签
    for each in targets_name:  # 剔除targets_name当中的别名
        if '\xa0' in each:
            targets_name.remove(each)
    return targets_name


def get_movie_top250_workers(soup):
    targets = soup.find_all('p', class_="")
    targets_workers = []
    for each in targets:
        targets_workers.append(
            each.text.replace('<p class="">', '').replace('\n                            ', '').replace('\xa0',
                                                                                                        '').replace(
                '\n                        ', ''))
    return targets_workers


def get_movie_top250_star(soup):
    targets = soup.find_all('div', class_="star")
    targets_star = re.findall(r'<span class="rating_num" property="v:average">(.*?)<\/span>', str(targets))
    return targets_star


def get_movie_top250_quote(soup):
    targets = soup.find_all('p', class_="quote")
    targets_quote = re.findall(r'<span class="inq">(.*?)<\/span>', str(targets))
    return targets_quote


def save_to_excel(name, workers, star, quote):
    wb = openpyxl.Workbook()
    ws = wb.active

    ws['A1'] = "电影名称"
    ws['B1'] = "工作人员"
    ws['C1'] = "评分"
    ws['D1'] = "描述"
    for i in range(len(name)):
        result = [name[i], workers[i], star[i], quote[i]]
        ws.append(result)
    wb.save("豆瓣电影TOP250.xlsx")

def autolabel(rects, ax, xpos='center'):  # 设置显示每一个条形图的值
        xpos = xpos.lower()  # normalize the case of the parameter
        ha = {'center': 'center', 'right': 'left', 'left': 'right'}
        offset = {'center': 0.5, 'right': 0.57, 'left': 0.43}  # x_txt = x + w*off
        for rect in rects:
            height = rect.get_height()
            ax.text(rect.get_x() + rect.get_width() * offset[xpos], 1.01 * height,
                    '{}'.format(height), ha=ha[xpos], va='bottom', size=6.8)
def Spe():
    numbers = 1
    name = []
    workers = []
    star = []
    quote = []
    result = []
    while numbers:
        url = 'https://movie.douban.com/top250?start={}&filter='.format(numbers - 1)
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.87 Safari/537.36'}
        res = requests.get(url, headers=headers)
        soup = BeautifulSoup(res.text, "html.parser")
        name_1 = get_movie_top250_name(soup)
        workers_1 = get_movie_top250_workers(soup)
        star_1 = get_movie_top250_star(soup)
        quote_1 = get_movie_top250_quote(soup)
        for i in range(len(name_1)-2):
            name.append(name_1[i])
            workers.append(workers_1[i])
            star.append(star_1[i])
            quote.append(quote_1[i])
            text.insert(END,str(i+1)+str(name[i])+'导演：'+str(workers[i])+'\n')
            text.insert(END, '开始时间：{}'.format(datetime.datetime.now()) + '\n')
            text.insert(END, '结束时间：{}'.format(datetime.datetime.now()) + '\n')


        numbers += 25
        if numbers > 250:
            break
    save_to_excel(name, workers, star, quote)
    for i in range(len(name_1)-1):
        cursor.execute(
            "insert into douban(name,worker,score) values (%s,%s,%s)",
            (name[i], workers[i], star[i])
        )


    ind = np.arange(len(star))
    fig, ax = plt.subplots()
    ax.set_xlabel('year')
    ax.set_ylabel('numbers')
    ax.set_title('Douban top 250 movie numbers by year')

    rext = ax.bar(ind, star, color='b', tick_label=star)
    autolabel(rext, ax)
    plt.xticks(np.arange(len(star)), rotation=-90, size=7.2)  # 设置X轴坐标的属性

    fig = plt.gcf()
    fig.set_size_inches(55.55,25)  # 设置图片大小
    plt.savefig('E:/爬虫数据/douban_year.png', dpi=200)  # 保存统计图到本地，必须在show（）方法前调用，否则得到的是一张空白图片,dpi是分辨率
    plt.show()
    plt.close()
    conn.commit()
    conn.close()



def start():
    th = threading.Thread(target=Spe)
    th.start()

def main():
    global url_input, text

    print('开始时间：{}'.format(datetime.datetime.now()))
    root = Tk()
    root.title('豆瓣电影')
    #窗口的大小，后面的加号是窗口在整个屏幕的位置
    root.geometry('750x600+398+279')
    #标签控件，窗口中放置文本组件
    Label(root,text='请输入下载的url:',font=("华文行楷",20),fg='black').grid(column=0, row=0)
    Label(root, text='请输入选择的线程数:', font=("华文行楷", 15), fg='black').grid(column=3, row=0)
    #定位 pack包 place位置 grid是网格式的布局

    #Entry是可输入文本框
    url_input=Entry(root,font=("微软雅黑",15))
    url_input.grid(row=0,column=1)

    #下拉列表
    comvalue = tkinter.StringVar()  # 窗体自带的文本，新建一个值
    comboxlist = ttk.Combobox(root, textvariable=comvalue)# 初始化
    comboxlist["values"] = ("1", "2", "3", "4")
    comboxlist.place(x=550,y=30)
    comboxlist.current(0)  # 选择第一个



    Label(root,text='豆瓣电影https://movie.douban.com',font=("微软雅黑",10),fg='black').grid(row=1)
    #列表控件
    text=Listbox(root,font=('微软雅黑',15),width=45,height=10)
    #columnspan 组件所跨越的列数
    text.grid(row=2,columnspan=2)
    #设置按钮 sticky对齐方式，N S W E
    button =Button(root,text='开始下载',font=("微软雅黑",15),command=start).grid(row=3,column=0,sticky=W)
    button =Button(root,text='退出',font=("微软雅黑",15),command=root.quit).grid(row=3,column=1,sticky=E)


    #使得窗口一直存在
    mainloop()

    print('结束时间：{}'.format(datetime.datetime.now()))


main()
